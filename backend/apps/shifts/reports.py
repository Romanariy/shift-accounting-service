from calendar import month_name
from decimal import Decimal

from django.db.models import Sum

from .constants import WorkType
from .models import CompanionEntry, Employee, ShiftEntry
from .payments import calculate_phone_amount, iter_month_dates


def format_money(value):
    return float(Decimal(value or 0))


def append_totals(ws, start_row, amount_column, employees):
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=1, value="Итоги")
    ws.cell(row=total_row + 1, column=1, value="Общая сумма за месяц")
    ws.cell(
        row=total_row + 1,
        column=2,
        value=f"=SUM({amount_column}{start_row}:{amount_column}{ws.max_row})",
    )

    row = total_row + 2
    for employee in employees:
        ws.cell(row=row, column=1, value=employee.short_name)
        ws.cell(
            row=row,
            column=2,
            value=f'=SUMIF(C{start_row}:C{total_row - 1},"{employee.short_name}",{amount_column}{start_row}:{amount_column}{total_row - 1})',
        )
        row += 1


def style_sheet(ws, widths):
    from openpyxl.styles import Alignment, Font, PatternFill

    header_fill = PatternFill("solid", fgColor="E2B714")
    header_font = Font(bold=True, color="111111")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def generate_month_report(year, month):
    from openpyxl import Workbook

    wb = Workbook()
    employees = list(Employee.objects.filter(is_active=True).order_by("sort_order", "short_name"))
    month_dates = list(iter_month_dates(year, month))
    start_date = month_dates[0]
    end_date = month_dates[-1]

    shifts_by_date = {}
    shifts = (
        ShiftEntry.objects.filter(date__range=(start_date, end_date), deleted_at__isnull=True)
        .select_related("employee")
        .order_by("date", "created_at")
    )
    for shift in shifts:
        shifts_by_date.setdefault(shift.date, []).append(shift)

    ws = wb.active
    ws.title = "Смены"
    ws.append(["Дата смены", "Кол-во часов", "Кто", "Роль", "Комментарий", "Сумма за день"])
    for current_date in month_dates:
        day_shifts = shifts_by_date.get(current_date)
        if not day_shifts:
            ws.append([current_date, 0, "Без админа", "", "", 0])
            continue

        for shift in day_shifts:
            ws.append(
                [
                    shift.date,
                    format_money(shift.hours),
                    shift.employee_name_snapshot,
                    shift.get_work_type_display(),
                    shift.comment,
                    format_money(shift.calculated_amount),
                ]
            )
    append_totals(ws, 2, "F", employees)
    style_sheet(ws, {"A": 14, "B": 14, "C": 18, "D": 24, "E": 34, "F": 16})

    companion_ws = wb.create_sheet("Сопровождения")
    companion_ws.append(["Дата", "Кол-во сопровождений", "Кто", "Сумма"])
    companions = (
        CompanionEntry.objects.filter(date__range=(start_date, end_date), deleted_at__isnull=True)
        .select_related("employee")
        .order_by("date", "created_at")
    )
    for companion in companions:
        companion_ws.append(
            [
                companion.date,
                companion.count,
                companion.employee_name_snapshot,
                format_money(companion.calculated_amount),
            ]
        )
    append_totals(companion_ws, 2, "D", employees)
    style_sheet(companion_ws, {"A": 14, "B": 24, "C": 18, "D": 16})

    phone_ws = wb.create_sheet("Телефоны")
    phone_ws.append(["Дата", "Сумма"])
    for current_date in month_dates:
        phone_ws.append([current_date, format_money(calculate_phone_amount(current_date))])
    total_row = phone_ws.max_row + 2
    phone_ws.cell(row=total_row, column=1, value="Итоги")
    phone_ws.cell(row=total_row + 1, column=1, value="Общая сумма за месяц")
    phone_ws.cell(row=total_row + 1, column=2, value=f"=SUM(B2:B{len(month_dates) + 1})")
    style_sheet(phone_ws, {"A": 14, "B": 16})

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):
            if row and getattr(row[0].value, "year", None):
                row[0].number_format = "DD.MM.YYYY"

    wb.properties.title = f"Отчет за {month_name[month]} {year}"
    return wb


def build_month_summary(year, month):
    month_dates = list(iter_month_dates(year, month))
    start_date = month_dates[0]
    end_date = month_dates[-1]
    employees = Employee.objects.filter(is_active=True).order_by("sort_order", "short_name")

    shift_total = (
        ShiftEntry.objects.filter(date__range=(start_date, end_date), deleted_at__isnull=True)
        .aggregate(total=Sum("calculated_amount"))
        .get("total")
        or 0
    )
    companion_total = (
        CompanionEntry.objects.filter(date__range=(start_date, end_date), deleted_at__isnull=True)
        .aggregate(total=Sum("calculated_amount"))
        .get("total")
        or 0
    )
    phone_total = sum(calculate_phone_amount(current_date) for current_date in month_dates)
    rows = []

    for employee in employees:
        shift_amount = (
            ShiftEntry.objects.filter(
                date__range=(start_date, end_date),
                employee=employee,
                deleted_at__isnull=True,
            )
            .aggregate(total=Sum("calculated_amount"))
            .get("total")
            or 0
        )
        companion_amount = (
            CompanionEntry.objects.filter(
                date__range=(start_date, end_date),
                employee=employee,
                deleted_at__isnull=True,
            )
            .aggregate(total=Sum("calculated_amount"))
            .get("total")
            or 0
        )
        rows.append(
            {
                "employeeId": employee.id,
                "employeeName": employee.short_name,
                "shiftAmount": str(shift_amount),
                "companionAmount": str(companion_amount),
                "totalAmount": str(Decimal(shift_amount) + Decimal(companion_amount)),
            }
        )

    return {
        "year": year,
        "month": month,
        "shiftTotal": str(shift_total),
        "companionTotal": str(companion_total),
        "phoneTotal": str(phone_total),
        "grandTotal": str(Decimal(shift_total) + Decimal(companion_total) + Decimal(phone_total)),
        "employees": rows,
    }
