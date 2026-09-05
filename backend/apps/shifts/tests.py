import json
import os
from pathlib import Path
import subprocess
import sys
from tempfile import TemporaryDirectory
from unittest import TestCase as UnitTestCase
from datetime import date, datetime, timezone as dt_timezone
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import patch

from django.test import SimpleTestCase, TestCase, override_settings
from openpyxl import load_workbook

from .ingest import ingest_telegram_message
from .models import AuditLog, CompanionEntry, Employee, Organization, PayRule, ShiftEntry, SyncOutbox
from .parser import ParseError, parse_shift_message
from .payments import calculate_shift_amount
from .reports import generate_month_report


class MigrationTests(UnitTestCase):
    def test_existing_data_migration(self):
        manage = Path(__file__).resolve().parents[2] / "manage.py"
        with TemporaryDirectory(prefix="shift-migration-") as directory:
            env = {**os.environ, "DJANGO_DB_ENGINE": "django.db.backends.sqlite3",
                   "DJANGO_DB_NAME": str(Path(directory) / "migration.sqlite3"), "SHIFT_SYNC_ENDPOINT": ""}

            def run(*args):
                result = subprocess.run([sys.executable, str(manage), *args], env=env, capture_output=True, text=True)
                self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

            run("migrate", "shifts", "0001", "--noinput")
            run("shell", "-c", """
from django.db import connection
from django.db.migrations.executor import MigrationExecutor
from datetime import date
apps = MigrationExecutor(connection).loader.project_state([('shifts', '0001_initial')]).apps
Employee = apps.get_model('shifts', 'Employee')
Shift = apps.get_model('shifts', 'ShiftEntry')
PayRule = apps.get_model('shifts', 'PayRule')
employee = Employee.objects.first()
employee.default_work_type = 'photobar'
employee.save()
Shift.objects.create(date=date(2026, 9, 6), employee=employee, work_type='photobar', hours=4, calculated_amount='999.00')
PayRule.objects.filter(code='small_admin').update(hourly_rate='275.00')
PayRule.objects.create(code='small_admin', title='Later', calculation_type='hourly', hourly_rate='300.00', active_from=date(2026, 10, 1), active_to=date(2026, 10, 31))
PayRule.objects.filter(code='cleaning').delete()
"""
            )
            run("migrate", "shifts", "0002", "--noinput")
            run("shell", "-c", """
from apps.shifts.models import Organization, Employee, ShiftEntry, PayRule, AuditLog
from decimal import Decimal
entry = ShiftEntry.objects.get()
assert entry.work_type == 'small_admin'
assert entry.calculated_amount == Decimal('999.00')
assert entry.organization_id is None
assert entry.employee.default_work_type == 'small_admin'
assert Organization.objects.count() == 3
assert not PayRule.objects.filter(code='photobar').exists()
assert AuditLog.objects.filter(entity_type='pay_rule', action='archive', before__code='photobar').exists()
for org in Organization.objects.all():
    rates = list(PayRule.objects.filter(organization=org, code='small_admin').order_by('active_from'))
    assert len(rates) == 2
    assert rates[0].hourly_rate == Decimal('275.00')
    assert rates[1].active_to.isoformat() == '2026-10-31'
    assert PayRule.objects.get(organization=org, code='cleaning').fixed_amount == Decimal('700.00')
assert PayRule.objects.filter(organization=None, code='small_admin').count() == 2
"""
            )


class ParserTests(SimpleTestCase):
    organizations = [SimpleNamespace(id=1, name="Фокус", aliases=[]),
                     SimpleNamespace(id=2, name="Фотобар", aliases=[]),
                     SimpleNamespace(id=3, name="Квин", aliases=["к"])]

    def parse(self, text, **kwargs):
        return parse_shift_message(text, organizations=self.organizations, aliases=["Наташа"], **kwargs)

    def test_optional_date_dot_time_night_alias_and_employee(self):
        today = date(2026, 9, 6)
        for text in ("10.00-16.00 фокус", "Наташа 10:00-16:00 ФОКУС.  "):
            with self.subTest(text=text):
                result = self.parse(text, today=today)
                self.assertEqual(result.date, today)
                self.assertEqual(result.hours, Decimal("6.00"))
                self.assertEqual(result.organization_id, 1)
        result = self.parse("01.04 Наташа 22:00–02:00 + 2 сопровождения к", today=today)
        self.assertEqual(result.date, date(2026, 4, 1))
        self.assertEqual(result.hours, Decimal("4.00"))
        self.assertEqual(result.employee_hint, "Наташа")
        self.assertEqual(result.organization_id, 3)
        self.assertEqual(result.companion_count, 2)
        self.assertEqual(result.comment, "")
        self.assertEqual(self.parse("01.04.25 уборка квин").date, date(2025, 4, 1))

    @override_settings(TIME_ZONE="Asia/Yekaterinburg")
    def test_local_midnight_and_year(self):
        with patch("django.utils.timezone.now", return_value=datetime(2026, 12, 31, 20, tzinfo=dt_timezone.utc)):
            self.assertEqual(self.parse("Уборка к").date, date(2027, 1, 1))
            self.assertEqual(self.parse("02.01 Уборка к").date, date(2027, 1, 2))

    def test_companion_only_and_work_types(self):
        self.assertFalse(self.parse("+ 2 сопр фотобар").has_shift)
        self.assertEqual(self.parse("Уборка фотобар").work_type, "cleaning")
        self.assertEqual(self.parse("Покраска циклораммы квин").work_type, "cyclorama_painting")
        self.assertEqual(self.parse("Большой админ квин").work_type, "big_admin")
        self.assertEqual(self.parse("10:00-16:00 фотобар").work_type, "default_shift")

    def test_invalid_messages(self):
        for text in ("", "фокус", "привет фокус", "10:00-16:00", "10:00-16:00 неизвестно",
                     "фокус 10:00-16:00", "10:00-16:00 антифокус", "31.02 уборка фокус",
                     "01.13 уборка к", "25:00-26:00 квин", "10:99-16:00 к", "9:0-12:00 к",
                     "100:00-16:00 к", "10:00-16:000 к", "+ 0 сопр к",
                     "+ 1 сопр + 2 сопр к", "10:00-12:00 + -1 сопр к",
                     "31/02 10:00-16:00 к", "10:00-12:00 13:00-15:00 к"):
            with self.subTest(text=text), self.assertRaises(ParseError):
                self.parse(text)


@override_settings(SHIFT_SYNC_AFTER_WRITE=False, SHIFT_SYNC_ENDPOINT="")
class OrganizationTests(TestCase):
    def setUp(self):
        self.focus = Organization.objects.get(name="Фокус")
        self.queen = Organization.objects.get(name="Квин")
        self.photobar = Organization.objects.get(name="Фотобар")
        self.employee = Employee.objects.get(short_name="Рамис")

    def api(self, method, url, payload):
        return getattr(self.client, method)("/api/shifts/" + url, json.dumps(payload), content_type="application/json")

    def entry_payload(self, **overrides):
        return {"date": "2026-09-06", "kind": "shift", "employeeId": self.employee.pk,
                "organizationId": self.focus.pk, "workType": "small_admin", "hours": "4", **overrides}

    def test_seed_rates_and_independent_organization_rates(self):
        for org in (self.focus, self.queen, self.photobar):
            for hours, expected in ((1, 600), (4, 800), (10, 1200)):
                self.assertEqual(calculate_shift_amount("small_admin", hours, date(2026, 9, 6), org.pk), expected)
            self.assertEqual(calculate_shift_amount("big_admin", 1, date(2026, 9, 6), org.pk), 1400)
        PayRule.objects.create(organization=self.queen, code="small_admin", title="Новая ставка",
                               calculation_type="hourly", hourly_rate=300, active_from=date(2026, 9, 1),
                               active_to=date(2026, 9, 30))
        self.assertEqual(calculate_shift_amount("small_admin", 4, date(2026, 9, 6), self.queen.pk), 1200)
        self.assertEqual(calculate_shift_amount("small_admin", 4, date(2026, 10, 1), self.queen.pk), 800)
        self.assertEqual(calculate_shift_amount("small_admin", 4, date(2026, 9, 6), self.focus.pk), 800)

    def test_bot_org_roles_companions_and_outbox(self):
        result = ingest_telegram_message("06.09 10:00-14:00 + 2 сопровождений фотобар", author_username=self.employee.telegram_username)
        self.assertEqual(result.shift.work_type, "small_admin")
        self.assertEqual(result.shift.calculated_amount, 800)
        self.assertEqual(result.companion.calculated_amount, 1000)
        self.assertEqual(result.companion.organization_id, self.photobar.pk)
        self.assertIn("Фотобар", result.message)
        self.assertEqual(SyncOutbox.objects.get(entity_type="shift", entity_id=result.shift.pk).payload["organizationId"], self.photobar.pk)
        count = ShiftEntry.objects.count()
        result = ingest_telegram_message("+ 2 сопр квин")
        self.assertIsNone(result.shift)
        self.assertEqual(ShiftEntry.objects.count(), count)
        self.assertTrue(result.needs_review)
        result = ingest_telegram_message("06.09 Наташа 10:00-12:00 фотобар")
        self.assertEqual(result.shift.work_type, "big_admin")
        self.assertEqual(result.shift.calculated_amount, 1400)

    def test_atomic_ingest_and_missing_tariff(self):
        with patch("apps.shifts.ingest.calculate_companion_amount", side_effect=ValueError("test")):
            with self.assertRaises(ValueError):
                ingest_telegram_message("06.09 10:00-14:00 + 2 сопр фокус")
        self.assertFalse(ShiftEntry.objects.exists())
        self.assertFalse(SyncOutbox.objects.exists())
        PayRule.objects.filter(organization=self.focus).update(is_active=False)
        with self.assertRaises(ParseError):
            ingest_telegram_message("06.09 10:00-14:00 фокус")
        response = self.api("post", "entries/", self.entry_payload())
        self.assertEqual(response.status_code, 400)
        self.assertFalse(ShiftEntry.objects.exists())

    def test_organization_crud_aliases_sheets_and_clone(self):
        response = self.api("post", "organizations/", {"name": "Студия", "aliases": ["ст"], "excelSheet": "Новый"})
        self.assertEqual(response.status_code, 201, response.content)
        org_id = response.json()["id"]
        self.assertEqual(PayRule.objects.filter(organization_id=org_id).count(), PayRule.objects.filter(organization=self.focus).count())
        for payload in ({"name": "Другая", "aliases": ["К"], "excelSheet": "Новый"},
                        {"name": "квин", "excelSheet": "Новый"},
                        {"name": "Другая", "excelSheet": "Телефоны"},
                        {"name": "Другая", "excelSheet": "Плохой/лист"},
                        {"name": "Другая", "excelSheet": "а" * 32}):
            self.assertEqual(self.api("post", "organizations/", payload).status_code, 400)
        response = self.api("put", f"organizations/{org_id}/", {"aliases": ["новая"], "excelSheet": "квин"})
        self.assertEqual(response.json()["excelSheet"], "Квин")
        self.assertEqual(ingest_telegram_message("Уборка новая").shift.organization_id, org_id)
        self.assertEqual(self.api("delete", f"organizations/{org_id}/", {}).status_code, 200)
        with self.assertRaises(ParseError):
            ingest_telegram_message("Уборка новая")
        self.assertTrue(AuditLog.objects.filter(entity_type="organization", entity_id=org_id).exists())

    def test_manual_required_legacy_and_inactive(self):
        self.assertEqual(self.api("post", "entries/", self.entry_payload(organizationId=None)).status_code, 400)
        self.assertEqual(self.api("post", "entries/", self.entry_payload(workType="photobar")).status_code, 400)
        result = self.api("post", "entries/", self.entry_payload())
        self.assertEqual(result.status_code, 201)
        self.assertEqual(result.json()["organizationName"], "Фокус")
        self.focus.is_active = False
        self.focus.save()
        self.assertEqual(self.api("post", "entries/", self.entry_payload()).status_code, 400)
        self.assertEqual(self.api("put", f"entries/shift/{result.json()['id']}/", self.entry_payload()).status_code, 200)
        legacy = ShiftEntry.objects.create(date=date(2026, 9, 6), work_type="small_admin", hours=4, calculated_amount=777)
        response = self.api("put", f"entries/shift/{legacy.pk}/", self.entry_payload(organizationId=None))
        self.assertEqual(response.status_code, 200, response.content)
        self.assertIsNone(response.json()["organizationId"])
        self.assertEqual(self.api("put", f"entries/shift/{legacy.pk}/", self.entry_payload(organizationId=self.queen.pk)).status_code, 200)

    def test_tariff_api_and_no_automatic_recalculation(self):
        entry = ingest_telegram_message("06.09 10:00-14:00 фокус", author_username=self.employee.telegram_username).shift
        rule = PayRule.objects.get(organization=self.focus, code="small_admin")
        payload = {"code": "small_admin", "organizationId": self.focus.pk, "title": "Малый",
                   "calculationType": "hourly", "hourlyRate": "500", "activeFrom": "2026-01-01"}
        response = self.api("put", f"pay-rules/{rule.pk}/", payload)
        self.assertEqual(response.status_code, 200, response.content)
        entry.refresh_from_db()
        self.assertEqual(entry.calculated_amount, 800)
        self.assertEqual(calculate_shift_amount("small_admin", 4, entry.date, self.focus.pk), 2000)
        self.assertEqual(self.api("post", "pay-rules/", {**payload, "organizationId": None}).status_code, 400)
        self.assertEqual(self.api("post", "pay-rules/", {**payload, "code": "companion"}).status_code, 400)
        self.assertEqual(self.api("post", "pay-rules/", {**payload, "hourlyRate": "-1"}).status_code, 400)

    def test_excel_groups_totals_and_disabled_history(self):
        for org in (self.focus, self.photobar, self.queen):
            ingest_telegram_message(f"06.09 10:00-14:00 + 1 сопр {org.name}", author_username=self.employee.telegram_username)
        ShiftEntry.objects.create(date=date(2026, 9, 6), employee=self.employee, work_type="small_admin", hours=4, calculated_amount=777)
        self.queen.is_active = False
        self.queen.save()
        wb = generate_month_report(2026, 9)
        output = BytesIO()
        wb.save(output)
        wb = load_workbook(BytesIO(output.getvalue()))
        self.assertEqual(wb.sheetnames, ["Фокус и Фотобар", "Квин", "Без организации", "Сопровождения", "Телефоны"])
        for sheet, count in (("Фокус и Фотобар", 2), ("Квин", 1), ("Без организации", 1)):
            rows = list(wb[sheet].values)
            self.assertEqual(sum(1 for row in rows if isinstance(row[0], datetime) and row[2] == "Рамис"), count)
            self.assertTrue(any(isinstance(row[1], str) and row[1].startswith("=SUM(F2:F") for row in rows))
            expected = 777 if sheet == "Без организации" else count * 800
            self.assertEqual(sum(row[5] for row in rows if isinstance(row[0], datetime)), expected)
        self.assertEqual(sum(1 for row in wb["Сопровождения"].values if isinstance(row[0], datetime)), 3)
        self.assertEqual(sum(row[3] for row in wb["Сопровождения"].values if isinstance(row[0], datetime)), 1500)
        self.photobar.excel_sheet = "Отдельный"
        self.photobar.save()
        self.assertIn("Отдельный", generate_month_report(2026, 9).sheetnames)
